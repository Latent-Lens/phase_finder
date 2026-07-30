#!/usr/bin/env python3
"""Build the FlowJo Dean-Jett-Fox (DJF) independent-reference set (VALID-01).

An analyst fitted FlowJo's Dean-Jett-Fox cell-cycle model to 30 asynchronous
budding-yeast samples and compiled the per-sample %G1/%S/%G2, G1/G2 means and
CVs, and the G2:G1 ratio into an Excel workbook. Those FCS files plus that
workbook are the direct DJF-vs-DJF external reference for PhaseFinder: unlike the
manual-gate / mass-cytometry references already in the manifest, this compares
the *same model family* run by an established tool, so close agreement (not just
"same direction") is the expectation.

This script is the traceable Excel-to-JSON mapping for that reference: it reads
the "Async ALL DATA" sheet, maps each strain to its source FCS file by the
`async_<strain>__` token in the filename, records file size + SHA-256 for
immutability, and writes datasets/flowjo_async_djf/flowjo_djf_reference.json.

This dataset is SPECIAL/PRIVATE: neither the FCS files nor this generated
reference JSON are committed. datasets/flowjo_async_djf/ is gitignored, and the
external validation manifest carries only a metadata stub (provenance, config,
tolerances, weight). Re-run this generator locally (pointing --xlsx / --fcs-dir
at the source data) whenever you need the reference JSON back.

Usage:
  /tmp/flowvenv/bin/python tests/validation/driving_code/generate_flowjo_djf_reference.py \
      [--xlsx PATH] [--fcs-dir DIR] [--no-hash] [--check]
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[3]
FLOW_PLOTTER = REPO_ROOT.parent  # the FlowPlotter workspace that holds test_flow_data/
DEFAULT_XLSX = FLOW_PLOTTER / "test_flow_data" / "DJF Model_DataCompilation_Flojo_3_26_2026.xlsx"
DEFAULT_WATSON_XLSX = FLOW_PLOTTER / "test_flow_data" / "DJF Model v. Watson Model.xlsx"
DEFAULT_FCS_DIR = FLOW_PLOTTER / "test_flow_data" / "Asynchronous_UsedAsFloJoDFJSampleDataset"
OUTPUT = REPO_ROOT / "tests/validation/validation_test_data/external_fcs/datasets/flowjo_async_djf/flowjo_djf_reference.json"

# FlowJo DJF workbook, "Async ALL DATA" column order (header verified 2026-07-28):
#   Strain, Count, G1%, S%, G2%, G1 Mean, G1 CV, G2 Mean, G2 CV, G2:G1 ratio
DJF_SHEET = "Async ALL DATA"
# Flowreader-Watson vs FlowJo-DJF workbook, "Sheet1" column order (row 3 header):
#   Strain, Count(W), Count(DJF), dCount, G1%(W), G1%(DJF), dG1, S%(W), S%(DJF),
#   dS, G2%(W), G2%(DJF), dG2, FSC-A Mean(W), FSC-A Mean(DJF), dFSC
WATSON_SHEET = "Sheet1"
STRAIN_RE = re.compile(r"async_([^_]+)__")


def strain_from_filename(name: str) -> str | None:
    match = STRAIN_RE.search(name)
    return match.group(1) if match else None


def sha256_of(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_flowjo_djf(xlsx: Path) -> dict[str, dict]:
    """strain -> FlowJo DJF reference (fractions, means, CVs, ratio, count)."""
    sheet = openpyxl.load_workbook(xlsx, data_only=True)[DJF_SHEET]
    out = {}
    for row in list(sheet.iter_rows(values_only=True))[1:]:
        strain, count = row[0], row[1]
        if not strain or count in (None, ""):
            continue
        out[str(strain).strip()] = {
            # FlowJo stores fractions (0.16) and CVs in percent (9.35). Keep the
            # analyst's exact values; consumers convert to whatever units they need.
            "count": int(count),
            "g1_fraction": float(row[2]),
            "s_fraction": float(row[3]),
            "g2_fraction": float(row[4]),
            "g1_mean": float(row[5]),
            "g1_cv_percent": float(row[6]),
            "g2_mean": float(row[7]),
            "g2_cv_percent": float(row[8]),
            "g2_g1_ratio": float(row[9]),
        }
    return out


def read_flowreader_watson(xlsx: Path, valid_strains: set[str]) -> dict[str, dict]:
    """strain -> Flowreader Watson reference (fractions + FSC-A mean) for the
    15-sample mat-a subset. Restricted to valid_strains so the averages/t-test
    rows below the data are ignored."""
    if not xlsx.exists():
        return {}
    sheet = openpyxl.load_workbook(xlsx, data_only=True)[WATSON_SHEET]
    out = {}
    for row in sheet.iter_rows(values_only=True):
        strain = row[0]
        if not strain or str(strain).strip() not in valid_strains:
            continue
        if row[1] in (None, "") or row[4] in (None, ""):
            continue
        out[str(strain).strip()] = {
            "count": int(row[1]),
            "g1_fraction": float(row[4]),
            "s_fraction": float(row[7]),
            "g2_fraction": float(row[10]),
            "fsc_a_mean": float(row[13]),
        }
    return out


def build_reference(xlsx: Path, watson_xlsx: Path, fcs_dir: Path, do_hash: bool) -> dict:
    djf = read_flowjo_djf(xlsx)
    watson = read_flowreader_watson(watson_xlsx, set(djf))
    fcs_by_strain = {}
    for path in sorted(fcs_dir.glob("*.fcs")):
        strain = strain_from_filename(path.name)
        if strain:
            fcs_by_strain[strain] = path

    records = []
    for strain in sorted(djf):
        path = fcs_by_strain.get(strain)
        fcs = None
        if path is not None:
            fcs = {"filename": path.name, "byte_size": path.stat().st_size}
            if do_hash:
                fcs["sha256"] = sha256_of(path)
        records.append({
            "strain": strain,
            "references": {
                "flowjo_djf": djf[strain],
                "flowreader_watson": watson.get(strain),
            },
            "fcs": fcs,
        })

    matched = sum(1 for r in records if r["fcs"] is not None)
    watson_count = sum(1 for r in records if r["references"]["flowreader_watson"] is not None)
    return {
        "schema_version": 2,
        "dataset_id": "flowjo_async_djf",
        "references_provided": {
            "flowjo_djf": {
                "software": "FlowJo",
                "model": "Dean-Jett-Fox (DJF)",
                "workbook": xlsx.name,
                "workbook_sheet": DJF_SHEET,
                "sample_count": len(records),
                "fields": ["count", "g1_fraction", "s_fraction", "g2_fraction", "g1_mean", "g1_cv_percent", "g2_mean", "g2_cv_percent", "g2_g1_ratio"],
            },
            "flowreader_watson": {
                "software": "Flowreader",
                "model": "Watson",
                "workbook": watson_xlsx.name,
                "workbook_sheet": WATSON_SHEET,
                "sample_count": watson_count,
                "subset": "mat-a strains only (1468*, 1693*, 1982*)",
                "fields": ["count", "g1_fraction", "s_fraction", "g2_fraction", "fsc_a_mean"],
                "caveat": "Flowreader runs the classic Watson model; PhaseFinder's 'Watson Pragmatic' restricts residual S to strictly between the fitted peaks (SCI-01). Flowreader's Watson S% here runs markedly higher than DJF S% (e.g. 1468f 0.41 vs 0.267), so PhaseFinder Watson Pragmatic is expected to sit LOWER than this reference. Compare broad agreement / direction, not strict equality.",
            },
        },
        "compilation_date": "2026-03-26",
        "generated_by": "tests/validation/driving_code/generate_flowjo_djf_reference.py",
        "organism": "budding yeast (asynchronous cultures)",
        "configuration_equivalence": {
            # Empirically confirmed 2026-07-28: FL7-A (GFP/FITC-A, SYTOX Green)
            # is the only channel whose 5th/95th percentiles (~160 / ~396 for
            # strain 1468f) bracket FlowJo's reported G1 mean (~175) and G2 mean
            # (~350). FL8-A (PI) sits far too low (median ~121). To reproduce the
            # FlowJo fit, PhaseFinder must model this channel.
            "dna_channel_pnn": "FL7-A",
            "dna_channel_pns": "GFP/FITC-A",
            "dna_stain": "SYTOX Green",
            "fcs_version": "3.1",
            "datatype": "F",
            "byte_order": "1,2,3,4",
            "parameters": 34,
            "channel_range_pnr": 1000,
            "note": "FlowJo's exact pre-fit gating (debris/singlet gates, live-cell scatter gate) is not captured in the workbook; residual gating differences are a documented source of small discrepancies.",
        },
        "sample_count": len(records),
        "fcs_matched": matched,
        "watson_matched": watson_count,
        "samples": records,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--watson-xlsx", type=Path, default=DEFAULT_WATSON_XLSX)
    parser.add_argument("--fcs-dir", type=Path, default=DEFAULT_FCS_DIR)
    parser.add_argument("--no-hash", action="store_true", help="skip SHA-256 (faster; sizes only)")
    parser.add_argument("--check", action="store_true", help="fail if the committed JSON is out of date")
    args = parser.parse_args()

    if not args.xlsx.exists():
        print(f"ERROR: workbook not found: {args.xlsx}")
        return 2
    reference = build_reference(args.xlsx, args.watson_xlsx, args.fcs_dir, do_hash=not args.no_hash)
    serialized = json.dumps(reference, indent=2) + "\n"

    if args.check:
        current = OUTPUT.read_text() if OUTPUT.exists() else ""
        # Sizes/hashes depend on local file presence; compare only the reference
        # science fields so --check is stable whether or not the FCS are present.
        def science(text):
            data = json.loads(text) if text else {}
            for s in data.get("samples", []):
                s.pop("fcs", None)
            data.pop("fcs_matched", None)
            data.pop("watson_matched", None)
            return data
        if science(serialized) != science(current):
            print("ERROR: committed flowjo_djf_reference.json is stale; re-run the generator.")
            return 1
        print("OK: reference science fields match the committed JSON.")
        return 0

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(serialized)
    print(f"Wrote {OUTPUT.relative_to(REPO_ROOT)}")
    print(f"  samples: {reference['sample_count']}, FCS matched: {reference['fcs_matched']}, hashed: {not args.no_hash}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
